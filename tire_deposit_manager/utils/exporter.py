#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Moduł do eksportu danych z bazy do różnych formatów plików.
"""

import os
import logging
import csv
import tempfile
from datetime import datetime

# Logger
logger = logging.getLogger("TireDepositManager")

def export_data_to_excel(conn, file_path, data_type, columns=None):
    """
    Eksportuje dane z bazy do pliku Excel.
    
    Args:
        conn: Połączenie z bazą danych
        file_path (str): Ścieżka do pliku docelowego
        data_type (str): Typ danych (np. 'clients', 'deposits', 'inventory')
        columns (list, optional): Lista kolumn do eksportu. Jeśli None, eksportuje wszystkie.
        
    Returns:
        int: Liczba wyeksportowanych rekordów
    """
    try:
        # Upewnijmy się, że mamy dostępne biblioteki
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("Brak biblioteki openpyxl. Zainstaluj ją: pip install openpyxl")
            raise ImportError("Brak biblioteki openpyxl. Zainstaluj ją: pip install openpyxl")
        
        # Przygotowanie zapytania SQL w zależności od typu danych
        cursor = conn.cursor()
        
        if data_type == "clients":
            # Eksport danych klientów
            table_name = "clients"
            all_columns = ["id", "name", "phone_number", "email", "additional_info", "discount", "barcode"]
            header_names = ["ID", "Nazwa klienta", "Telefon", "E-mail", "Informacje dodatkowe", "Rabat (%)", "Kod kreskowy"]
            query_template = "SELECT {columns} FROM clients"
            
            # Dodatkowe zapytanie o liczbie depozytów
            if columns and "deposit_count" in columns:
                query_template = """
                    SELECT c.{columns}, 
                    COUNT(d.id) AS deposit_count
                    FROM clients c
                    LEFT JOIN deposits d ON c.id = d.client_id
                    GROUP BY c.id
                """
                all_columns.append("deposit_count")
                header_names.append("Liczba depozytów")
        
        elif data_type == "deposits":
            # Eksport danych depozytów
            table_name = "deposits"
            all_columns = [
                "d.id", "c.name AS client_name", "d.car_model", "d.registration_number", 
                "d.tire_brand", "d.tire_size", "d.quantity", "d.location", 
                "d.deposit_date", "d.expected_return_date", "d.status", "d.season", "d.price"
            ]
            header_names = [
                "ID", "Klient", "Model auta", "Nr rejestracyjny", "Marka opon", 
                "Rozmiar opon", "Ilość", "Lokalizacja", "Data depozytu", 
                "Oczekiwany zwrot", "Status", "Sezon", "Cena (PLN)"
            ]
            query_template = """
                SELECT {columns}
                FROM deposits d
                LEFT JOIN clients c ON d.client_id = c.id
            """
        
        elif data_type == "inventory":
            # Eksport danych magazynu opon
            table_name = "inventory"
            all_columns = [
                "id", "brand_model", "size", "quantity", "price", 
                "dot", "season_type", "notes"
            ]
            header_names = [
                "ID", "Marka i model", "Rozmiar", "Ilość", "Cena (PLN)", 
                "DOT", "Sezon", "Uwagi"
            ]
            query_template = "SELECT {columns} FROM inventory"
        
        elif data_type == "parts":
            # Eksport danych części i akcesoriów
            table_name = "parts"
            all_columns = [
                "id", "name", "catalog_number", "category", "manufacturer", 
                "quantity", "price", "location", "description", "barcode", "minimum_quantity"
            ]
            header_names = [
                "ID", "Nazwa", "Nr katalogowy", "Kategoria", "Producent", 
                "Ilość", "Cena (PLN)", "Lokalizacja", "Opis", "Kod kreskowy", "Minimalna ilość"
            ]
            query_template = "SELECT {columns} FROM parts"
        
        elif data_type == "appointments":
            # Eksport danych wizyt
            table_name = "appointments"
            all_columns = [
                "a.id", "c.name AS client_name", "a.appointment_date", "a.appointment_time", 
                "a.service_type", "a.status", "a.notes", "a.duration"
            ]
            header_names = [
                "ID", "Klient", "Data", "Godzina", "Usługa", 
                "Status", "Uwagi", "Czas trwania (min)"
            ]
            query_template = """
                SELECT {columns}
                FROM appointments a
                LEFT JOIN clients c ON a.client_id = c.id
            """
        
        else:
            logger.error(f"Nieznany typ danych: {data_type}")
            return 0
        
        # Filtruj kolumny, jeśli podano konkretne
        if columns:
            # Wybierz tylko te kolumny, które są w all_columns
            selected_columns = []
            selected_headers = []
            for i, col in enumerate(all_columns):
                # Wyczyść nazwę kolumny do porównania
                clean_col = col.split(' AS ')[0].split('.')[-1]
                if clean_col in columns:
                    selected_columns.append(col)
                    selected_headers.append(header_names[i])
            
            columns_str = ", ".join(selected_columns)
            header_row = selected_headers
        else:
            columns_str = ", ".join(all_columns)
            header_row = header_names
        
        # Wykonaj zapytanie
        query = query_template.format(columns=columns_str)
        cursor.execute(query)
        rows = cursor.fetchall()
        
        # Jeśli nie ma danych, zwróć 0
        if not rows:
            logger.warning(f"Brak danych do eksportu typu: {data_type}")
            return 0
        
        # Stwórz nowy plik Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = data_type.capitalize()
        
        # Dodaj nagłówki
        for col_idx, header in enumerate(header_row, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1F618D", end_color="1F618D", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dodaj dane
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, cell_data in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                if cell_data is None:
                    cell.value = ""
                else:
                    # Formatowanie daty jeśli to pole daty
                    if any(date_field in header_row[col_idx-1].lower() for date_field in ["data", "date"]):
                        try:
                            # Sprawdź czy to data w formacie "yyyy-mm-dd"
                            if isinstance(cell_data, str) and len(cell_data) >= 10 and cell_data[4] == '-' and cell_data[7] == '-':
                                date_obj = datetime.strptime(cell_data[:10], "%Y-%m-%d")
                                cell.value = date_obj
                                cell.number_format = "yyyy-mm-dd"
                            else:
                                cell.value = cell_data
                        except (ValueError, TypeError):
                            cell.value = cell_data
                    
                    # Formatowanie ceny jeśli to pole ceny
                    elif any(price_field in header_row[col_idx-1].lower() for price_field in ["cena", "price"]):
                        try:
                            if isinstance(cell_data, (int, float)):
                                cell.value = float(cell_data)
                                cell.number_format = "#,##0.00 PLN"
                            else:
                                float_val = float(str(cell_data).replace(' PLN', '').replace(',', '.'))
                                cell.value = float_val
                                cell.number_format = "#,##0.00 PLN"
                        except (ValueError, TypeError):
                            cell.value = cell_data
                    
                    # Standardowa wartość dla innych typów
                    else:
                        cell.value = cell_data
        
        # Auto-formatowanie szerokości kolumn
        for col_idx in range(1, len(header_row) + 1):
            col_letter = get_column_letter(col_idx)
            # Ustaw szerokość kolumny na szerokość najdłuższej wartości + margines
            max_length = 0
            for row_idx in range(1, len(rows) + 2):  # +2 bo nagłówek + 1 (indeksujemy od 1)
                cell = worksheet[f"{col_letter}{row_idx}"]
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[col_letter].width = max(max_length + 2, 10)
        
        # Dodaj obramowania
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(rows) + 1, min_col=1, max_col=len(header_row)):
            for cell in row:
                cell.border = thin_border
        
        # Dodaj zebrowe pasy dla lepszej czytelności
        for row_idx in range(2, len(rows) + 2):
            if row_idx % 2 == 0:  # Parzyste wiersze
                for col_idx in range(1, len(header_row) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        
        # Zamroź pierwszy wiersz (nagłówki)
        worksheet.freeze_panes = "A2"
        
        # Dodaj filtrowanie
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(header_row))}{len(rows) + 1}"
        
        # Dodaj stopkę z datą eksportu i liczbą rekordów
        footer_row = len(rows) + 3
        cell = worksheet.cell(row=footer_row, column=1)
        cell.value = f"Data eksportu: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        cell.font = Font(italic=True)
        
        cell = worksheet.cell(row=footer_row, column=len(header_row))
        cell.value = f"Liczba rekordów: {len(rows)}"
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal="right")
        
        # Zapisz plik
        workbook.save(file_path)
        
        logger.info(f"Wyeksportowano {len(rows)} rekordów do pliku Excel: {file_path}")
        return len(rows)
    
    except Exception as e:
        logger.error(f"Błąd podczas eksportu do Excel: {e}")
        raise

def export_data_to_pdf(conn, file_path, data_type, columns=None, template=None):
    """
    Eksportuje dane z bazy do pliku PDF.
    
    Args:
        conn: Połączenie z bazą danych
        file_path (str): Ścieżka do pliku docelowego
        data_type (str): Typ danych (np. 'clients', 'deposits', 'inventory')
        columns (list, optional): Lista kolumn do eksportu. Jeśli None, eksportuje wszystkie.
        template (str, optional): Szablon do formatowania pliku PDF
        
    Returns:
        int: Liczba wyeksportowanych rekordów
    """
    try:
        # Upewnijmy się, że mamy dostępne biblioteki
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from utils.paths import RESOURCES_DIR, FONTS_DIR
        except ImportError:
            logger.error("Brak biblioteki reportlab. Zainstaluj ją: pip install reportlab")
            raise ImportError("Brak biblioteki reportlab. Zainstaluj ją: pip install reportlab")
        
        # Zarejestruj czcionki (opcjonalnie)
        try:
            font_path = os.path.join(FONTS_DIR, "DejaVuSans.ttf")
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
                font_name = 'DejaVuSans'
            else:
                font_name = 'Helvetica'
        except Exception as e:
            logger.warning(f"Nie można zarejestrować niestandardowej czcionki: {e}")
            font_name = 'Helvetica'
        
        # Określenie rozmiaru strony
        page_size = landscape(A4) if len(columns or []) > 5 else A4
        
        # Przygotowanie danych i nagłówków, tak jak w funkcji export_to_excel
        cursor = conn.cursor()
        
        # Ta sama logika co w export_to_excel - przygotowanie zapytania
        if data_type == "clients":
            # Eksport danych klientów
            table_name = "clients"
            all_columns = ["id", "name", "phone_number", "email", "additional_info", "discount", "barcode"]
            header_names = ["ID", "Nazwa klienta", "Telefon", "E-mail", "Informacje dodatkowe", "Rabat (%)", "Kod kreskowy"]
            query_template = "SELECT {columns} FROM clients"
            
            if columns and "deposit_count" in columns:
                query_template = """
                    SELECT c.{columns}, 
                    COUNT(d.id) AS deposit_count
                    FROM clients c
                    LEFT JOIN deposits d ON c.id = d.client_id
                    GROUP BY c.id
                """
                all_columns.append("deposit_count")
                header_names.append("Liczba depozytów")
            
            report_title = "Lista klientów"
        
        elif data_type == "deposits":
            # Eksport danych depozytów
            table_name = "deposits"
            all_columns = [
                "d.id", "c.name AS client_name", "d.car_model", "d.registration_number", 
                "d.tire_brand", "d.tire_size", "d.quantity", "d.location", 
                "d.deposit_date", "d.expected_return_date", "d.status", "d.season", "d.price"
            ]
            header_names = [
                "ID", "Klient", "Model auta", "Nr rejestracyjny", "Marka opon", 
                "Rozmiar opon", "Ilość", "Lokalizacja", "Data depozytu", 
                "Oczekiwany zwrot", "Status", "Sezon", "Cena (PLN)"
            ]
            query_template = """
                SELECT {columns}
                FROM deposits d
                LEFT JOIN clients c ON d.client_id = c.id
            """
            report_title = "Lista depozytów"
        
        elif data_type == "inventory":
            # Eksport danych magazynu opon
            table_name = "inventory"
            all_columns = [
                "id", "brand_model", "size", "quantity", "price", 
                "dot", "season_type", "notes"
            ]
            header_names = [
                "ID", "Marka i model", "Rozmiar", "Ilość", "Cena (PLN)", 
                "DOT", "Sezon", "Uwagi"
            ]
            query_template = "SELECT {columns} FROM inventory"
            report_title = "Stan magazynowy opon"
        
        elif data_type == "parts":
            # Eksport danych części i akcesoriów
            table_name = "parts"
            all_columns = [
                "id", "name", "catalog_number", "category", "manufacturer", 
                "quantity", "price", "location", "description", "barcode", "minimum_quantity"
            ]
            header_names = [
                "ID", "Nazwa", "Nr katalogowy", "Kategoria", "Producent", 
                "Ilość", "Cena (PLN)", "Lokalizacja", "Opis", "Kod kreskowy", "Minimalna ilość"
            ]
            query_template = "SELECT {columns} FROM parts"
            report_title = "Lista części i akcesoriów"
        
        elif data_type == "appointments":
            # Eksport danych wizyt
            table_name = "appointments"
            all_columns = [
                "a.id", "c.name AS client_name", "a.appointment_date", "a.appointment_time", 
                "a.service_type", "a.status", "a.notes", "a.duration"
            ]
            header_names = [
                "ID", "Klient", "Data", "Godzina", "Usługa", 
                "Status", "Uwagi", "Czas trwania (min)"
            ]
            query_template = """
                SELECT {columns}
                FROM appointments a
                LEFT JOIN clients c ON a.client_id = c.id
            """
            report_title = "Harmonogram wizyt"
        
        else:
            logger.error(f"Nieznany typ danych: {data_type}")
            return 0
        
        # Filtruj kolumny, jeśli podano konkretne
        if columns:
            selected_columns = []
            selected_headers = []
            for i, col in enumerate(all_columns):
                clean_col = col.split(' AS ')[0].split('.')[-1]
                if clean_col in columns:
                    selected_columns.append(col)
                    selected_headers.append(header_names[i])
            
            columns_str = ", ".join(selected_columns)
            header_row = selected_headers
        else:
            columns_str = ", ".join(all_columns)
            header_row = header_names
        
        # Wykonaj zapytanie
        query = query_template.format(columns=columns_str)
        cursor.execute(query)
        rows = cursor.fetchall()
        
        # Jeśli nie ma danych, zwróć 0
        if not rows:
            logger.warning(f"Brak danych do eksportu typu: {data_type}")
            return 0
        

        # Dodaj własny styl nagłówka
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=14,
            spaceAfter=12,
            alignment=1  # Wyśrodkowanie
        )

        # Stwórz styl dla komórek
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8
        )

        # Stwórz styl dla stopki
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8,
            textColor=colors.gray
        )

# Pobierz style dla dokumentu
        styles = getSampleStyleSheet()
        
        # Dodaj własny styl nagłówka
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Heading1'],
            fontName=font_name,
            fontSize=14,
            spaceAfter=12,
            alignment=1  # Wyśrodkowanie
        )
        
        # Stwórz styl dla komórek
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8
        )
        
        # Stwórz styl dla stopki
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=8,
            textColor=colors.gray
        )
        
        # Utwórz dokument PDF
        doc = SimpleDocTemplate(
            file_path,
            pagesize=page_size,
            leftMargin=1*cm,
            rightMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        
        # Lista elementów do dodania do dokumentu
        elements = []
        
        # Dodaj logo i tytuł, jeśli jest dostępny szablon
        if template and template == "z_logo":
            # Dodaj logo
            logo_path = os.path.join(RESOURCES_DIR, "images", "logo.png")
            if os.path.exists(logo_path):
                logo = Image(logo_path)
                logo.drawHeight = 1.5*cm
                logo.drawWidth = 5*cm
                elements.append(logo)
                elements.append(Spacer(1, 0.5*cm))
        
        # Dodaj tytuł
        title = Paragraph(f"{report_title}", header_style)
        elements.append(title)
        elements.append(Spacer(1, 0.5*cm))
        
        # Przygotuj dane tabeli
        table_data = [header_row]
        
        # Dodaj dane do tabeli, formatując je odpowiednio
        for row_data in rows:
            formatted_row = []
            for i, cell in enumerate(row_data):
                if cell is None:
                    formatted_row.append("")
                elif any(date_field in header_row[i].lower() for date_field in ["data", "date"]):
                    # Formatowanie daty
                    try:
                        if isinstance(cell, str) and len(cell) >= 10 and cell[4] == '-' and cell[7] == '-':
                            date_obj = datetime.strptime(cell[:10], "%Y-%m-%d")
                            formatted_row.append(date_obj.strftime("%d.%m.%Y"))
                        else:
                            formatted_row.append(str(cell))
                    except (ValueError, TypeError):
                        formatted_row.append(str(cell))
                elif any(price_field in header_row[i].lower() for price_field in ["cena", "price"]):
                    # Formatowanie ceny
                    try:
                        price = float(str(cell).replace(' PLN', '').replace(',', '.'))
                        formatted_row.append(f"{price:.2f} PLN")
                    except (ValueError, TypeError):
                        formatted_row.append(str(cell))
                else:
                    formatted_row.append(str(cell))
            
            table_data.append(formatted_row)
        
        # Tworzenie tabeli
        table = Table(table_data, repeatRows=1)
        
        # Styl tabeli
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F618D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ])
        
        # Dodaj zebrowe pasy dla lepszej czytelności
        for i in range(1, len(rows) + 1):
            if i % 2 == 0:  # Parzyste wiersze
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#ECF0F1'))
        
        table.setStyle(table_style)
        
        # Dodaj tabelę do dokumentu
        elements.append(table)
        
        # Dodaj stopkę
        elements.append(Spacer(1, 1*cm))
        footer_text = f"Data wygenerowania: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Liczba rekordów: {len(rows)}"
        footer = Paragraph(footer_text, footer_style)
        elements.append(footer)
        
        # Zbuduj dokument
        doc.build(elements)
        
        logger.info(f"Wyeksportowano {len(rows)} rekordów do pliku PDF: {file_path}")
        return len(rows)
    
    except Exception as e:
        logger.error(f"Błąd podczas eksportu do PDF: {e}")
        raise

def export_data_to_csv(conn, file_path, data_type, columns=None):
    """
    Eksportuje dane z bazy do pliku CSV.
    
    Args:
        conn: Połączenie z bazą danych
        file_path (str): Ścieżka do pliku docelowego
        data_type (str): Typ danych (np. 'clients', 'deposits', 'inventory')
        columns (list, optional): Lista kolumn do eksportu. Jeśli None, eksportuje wszystkie.
        
    Returns:
        int: Liczba wyeksportowanych rekordów
    """
    try:
        # Przygotowanie zapytania SQL w zależności od typu danych
        cursor = conn.cursor()
        
        if data_type == "clients":
            # Eksport danych klientów
            table_name = "clients"
            all_columns = ["id", "name", "phone_number", "email", "additional_info", "discount", "barcode"]
            header_names = ["ID", "Nazwa klienta", "Telefon", "E-mail", "Informacje dodatkowe", "Rabat (%)", "Kod kreskowy"]
            query_template = "SELECT {columns} FROM clients"
            
            # Dodatkowe zapytanie o liczbie depozytów
            if columns and "deposit_count" in columns:
                query_template = """
                    SELECT c.{columns}, 
                    COUNT(d.id) AS deposit_count
                    FROM clients c
                    LEFT JOIN deposits d ON c.id = d.client_id
                    GROUP BY c.id
                """
                all_columns.append("deposit_count")
                header_names.append("Liczba depozytów")
        
        elif data_type == "deposits":
            # Eksport danych depozytów
            table_name = "deposits"
            all_columns = [
                "d.id", "c.name AS client_name", "d.car_model", "d.registration_number", 
                "d.tire_brand", "d.tire_size", "d.quantity", "d.location", 
                "d.deposit_date", "d.expected_return_date", "d.status", "d.season", "d.price"
            ]
            header_names = [
                "ID", "Klient", "Model auta", "Nr rejestracyjny", "Marka opon", 
                "Rozmiar opon", "Ilość", "Lokalizacja", "Data depozytu", 
                "Oczekiwany zwrot", "Status", "Sezon", "Cena (PLN)"
            ]
            query_template = """
                SELECT {columns}
                FROM deposits d
                LEFT JOIN clients c ON d.client_id = c.id
            """
        
        elif data_type == "inventory":
            # Eksport danych magazynu opon
            table_name = "inventory"
            all_columns = [
                "id", "brand_model", "size", "quantity", "price", 
                "dot", "season_type", "notes"
            ]
            header_names = [
                "ID", "Marka i model", "Rozmiar", "Ilość", "Cena (PLN)", 
                "DOT", "Sezon", "Uwagi"
            ]
            query_template = "SELECT {columns} FROM inventory"
        
        elif data_type == "parts":
            # Eksport danych części i akcesoriów
            table_name = "parts"
            all_columns = [
                "id", "name", "catalog_number", "category", "manufacturer", 
                "quantity", "price", "location", "description", "barcode", "minimum_quantity"
            ]
            header_names = [
                "ID", "Nazwa", "Nr katalogowy", "Kategoria", "Producent", 
                "Ilość", "Cena (PLN)", "Lokalizacja", "Opis", "Kod kreskowy", "Minimalna ilość"
            ]
            query_template = "SELECT {columns} FROM parts"
        
        elif data_type == "appointments":
            # Eksport danych wizyt
            table_name = "appointments"
            all_columns = [
                "a.id", "c.name AS client_name", "a.appointment_date", "a.appointment_time", 
                "a.service_type", "a.status", "a.notes", "a.duration"
            ]
            header_names = [
                "ID", "Klient", "Data", "Godzina", "Usługa", 
                "Status", "Uwagi", "Czas trwania (min)"
            ]
            query_template = """
                SELECT {columns}
                FROM appointments a
                LEFT JOIN clients c ON a.client_id = c.id
            """
        
        else:
            logger.error(f"Nieznany typ danych: {data_type}")
            return 0
        
        # Filtruj kolumny, jeśli podano konkretne
        if columns:
            # Wybierz tylko te kolumny, które są w all_columns
            selected_columns = []
            selected_headers = []
            for i, col in enumerate(all_columns):
                # Wyczyść nazwę kolumny do porównania
                clean_col = col.split(' AS ')[0].split('.')[-1]
                if clean_col in columns:
                    selected_columns.append(col)
                    selected_headers.append(header_names[i])
            
            columns_str = ", ".join(selected_columns)
            header_row = selected_headers
        else:
            columns_str = ", ".join(all_columns)
            header_row = header_names
        
        # Wykonaj zapytanie
        query = query_template.format(columns=columns_str)
        cursor.execute(query)
        rows = cursor.fetchall()
        
        # Jeśli nie ma danych, zwróć 0
        if not rows:
            logger.warning(f"Brak danych do eksportu typu: {data_type}")
            return 0
        
        # Otwórz plik CSV do zapisu
        with open(file_path, 'w', newline='', encoding='utf-8-sig') as csv_file:
            csv_writer = csv.writer(csv_file, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            
            # Zapisz nagłówki
            csv_writer.writerow(header_row)
            
            # Zapisz dane
            for row in rows:
                # Przekształcenie None na puste stringi i formatowanie danych
                formatted_row = []
                for i, cell in enumerate(row):
                    if cell is None:
                        formatted_row.append("")
                    elif any(date_field in header_row[i].lower() for date_field in ["data", "date"]):
                        # Formatowanie daty
                        try:
                            if isinstance(cell, str) and len(cell) >= 10 and cell[4] == '-' and cell[7] == '-':
                                date_obj = datetime.strptime(cell[:10], "%Y-%m-%d")
                                formatted_row.append(date_obj.strftime("%d.%m.%Y"))
                            else:
                                formatted_row.append(str(cell))
                        except (ValueError, TypeError):
                            formatted_row.append(str(cell))
                    elif any(price_field in header_row[i].lower() for price_field in ["cena", "price"]):
                        # Formatowanie ceny
                        try:
                            price = float(str(cell).replace(' PLN', '').replace(',', '.'))
                            formatted_row.append(f"{price:.2f} PLN".replace('.', ','))
                        except (ValueError, TypeError):
                            formatted_row.append(str(cell))
                    else:
                        formatted_row.append(str(cell))
                
                csv_writer.writerow(formatted_row)
        
        logger.info(f"Wyeksportowano {len(rows)} rekordów do pliku CSV: {file_path}")
        return len(rows)
    
    except Exception as e:
        logger.error(f"Błąd podczas eksportu do CSV: {e}")
        raise